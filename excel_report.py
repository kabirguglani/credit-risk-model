import warnings
warnings.filterwarnings('ignore')

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, Reference
import os

BASE = '/Users/kabir/Desktop/pro/credit_risk/'

# ── Load data ──────────────────────────────────────────────────────────────────
print("Loading data...")
portfolio  = pd.read_csv(BASE + 'final_portfolio.csv')
stress     = pd.read_csv(BASE + 'stress_test_results.csv')
migration  = pd.read_csv(BASE + 'grade_migration.csv', index_col=0)

# ── Colours ────────────────────────────────────────────────────────────────────
NAVY_HEX   = "1F3864"
BLUE_HEX   = "2E5FA3"
GOLD_HEX   = "C9993A"
RED_HEX    = "C0392B"
GREEN_HEX  = "1A7A4A"
LGREY_HEX  = "E8EEF7"
WHITE_HEX  = "FFFFFF"
MGREY_HEX  = "F7F9FC"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def border_thin():
    s = Side(style='thin', color='D1D9E6')
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── Create workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: Executive Summary
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 20
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 20
ws1.column_dimensions['F'].width = 20
ws1.column_dimensions['G'].width = 3

# Title banner
for row in range(1, 5):
    for col in range(1, 8):
        ws1.cell(row=row, column=col).fill = fill(NAVY_HEX)

ws1.merge_cells('B2:F2')
ws1['B2'] = 'CREDIT RISK & STRESS TESTING MODEL'
ws1['B2'].font = font(bold=True, color=WHITE_HEX, size=16)
ws1['B2'].alignment = center()

ws1.merge_cells('B3:F3')
ws1['B3'] = 'German Credit Dataset · Basel II Framework · PD/LGD/EAD Analysis'
ws1['B3'].font = font(color='C8D8F0', size=10, italic=True)
ws1['B3'].alignment = center()

ws1.merge_cells('B4:F4')
ws1['B4'] = 'Kabir Guglani | MS Financial Analysis | Temple University | 2025'
ws1['B4'].font = font(color='A0B8D8', size=9)
ws1['B4'].alignment = center()

# KPI boxes row
ws1.row_dimensions[6].height = 50
kpis = [
    ('1,000', 'Loans Analysed'),
    ('30.0%', 'Portfolio Default Rate'),
    ('DM 3.27M', 'Total Exposure'),
    ('0.833', 'Model AUC-ROC'),
    ('0.666', 'Gini Coefficient'),
]
kpi_cols = ['B','C','D','E','F']
for col_letter, (val, label) in zip(kpi_cols, kpis):
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    cell_val = ws1.cell(row=6, column=col_idx)
    cell_lbl = ws1.cell(row=7, column=col_idx)
    cell_val.value = val
    cell_val.font = font(bold=True, color=NAVY_HEX, size=14)
    cell_val.alignment = center()
    cell_val.fill = fill(LGREY_HEX)
    cell_lbl.value = label
    cell_lbl.font = font(color='6B7280', size=9)
    cell_lbl.alignment = center()
    cell_lbl.fill = fill(LGREY_HEX)

# Section: Base Case
ws1.row_dimensions[9].height = 20
ws1.merge_cells('B9:F9')
ws1['B9'] = 'BASE CASE — Portfolio Summary'
ws1['B9'].font = font(bold=True, color=WHITE_HEX, size=11)
ws1['B9'].fill = fill(BLUE_HEX)
ws1['B9'].alignment = left()

base_data = [
    ('Average PD', '44.2%', 'Average EL Rate', '23.7%'),
    ('Total Expected Loss', 'DM 862,226', 'EL as % of Portfolio', '26.4%'),
    ('Tier 1 Capital (8%)', 'DM 261,701', 'Capital Adequacy', '❌ Requires buffer'),
]
for i, (l1, v1, l2, v2) in enumerate(base_data):
    r = 10 + i
    ws1.row_dimensions[r].height = 18
    bg = MGREY_HEX if i % 2 == 0 else WHITE_HEX
    for col, val, bold in [(2,l1,True),(3,v1,False),(4,l2,True),(5,v2,False)]:
        c = ws1.cell(row=r, column=col)
        c.value = val; c.font = font(bold=bold, size=10)
        c.fill = fill(bg); c.border = border_thin()
        c.alignment = left() if col in [2,4] else center()

# Section: Stress scenarios
ws1.row_dimensions[14].height = 20
ws1.merge_cells('B14:F14')
ws1['B14'] = 'STRESS TEST RESULTS'
ws1['B14'].font = font(bold=True, color=WHITE_HEX, size=11)
ws1['B14'].fill = fill(NAVY_HEX)
ws1['B14'].alignment = left()

headers = ['Scenario', 'Avg PD', 'Total EL (DM)', 'EL % Portfolio', 'vs Base']
for j, h in enumerate(headers):
    c = ws1.cell(row=15, column=2+j)
    c.value = h; c.font = font(bold=True, color=WHITE_HEX, size=9)
    c.fill = fill(BLUE_HEX); c.alignment = center(); c.border = border_thin()

scenario_colors = [MGREY_HEX, 'FFE8E8', 'FFF5E0', 'F0E8FF', 'FFF0E0']
stress_display = [
    ('Base Case',             '44.2%', 'DM 862,226',   '26.4%', '—'),
    ('2008 GFC',              '76.0%', 'DM 1,805,278', '55.2%', '+109%'),
    ('2020 COVID',            '66.0%', 'DM 1,422,734', '43.5%', '+65%'),
    ('Rising Rates (+300bps)','58.4%', 'DM 1,225,613', '37.5%', '+42%'),
    ('Mild Recession',        '60.9%', 'DM 1,270,345', '38.8%', '+47%'),
]
for i, (row_data, bg) in enumerate(zip(stress_display, scenario_colors)):
    r = 16 + i
    ws1.row_dimensions[r].height = 18
    for j, val in enumerate(row_data):
        c = ws1.cell(row=r, column=2+j)
        c.value = val
        c.font = font(bold=(j==0), size=9,
                      color=RED_HEX if (j==4 and val not in ['—','+0%']) else '000000')
        c.fill = fill(bg); c.border = border_thin()
        c.alignment = left() if j == 0 else center()

# Section: Key findings
ws1.row_dimensions[22].height = 20
ws1.merge_cells('B22:F22')
ws1['B22'] = 'KEY FINDINGS'
ws1['B22'].font = font(bold=True, color=WHITE_HEX, size=11)
ws1['B22'].fill = fill(NAVY_HEX)
ws1['B22'].alignment = left()

findings = [
    '✅  Model AUC 0.833 / Gini 0.666 — well above industry benchmark of 0.70',
    '✅  Checking account status is strongest predictor (IV = 0.666 — Very Strong)',
    '⚠️  GFC scenario increases Expected Loss by 109% — DM 943K additional loss',
    '⚠️  Under GFC, CCC-rated loans increase from 462 to 792 (+71% grade migration)',
    '❌  Capital buffer insufficient under all scenarios — requires additional provisioning',
    '📋  Recommended: maintain economic capital buffer of 3.5x regulatory minimum',
]
for i, finding in enumerate(findings):
    r = 23 + i
    ws1.row_dimensions[r].height = 18
    ws1.merge_cells(f'B{r}:F{r}')
    c = ws1[f'B{r}']
    c.value = finding
    c.font = font(size=9, color=NAVY_HEX)
    c.fill = fill(MGREY_HEX if i % 2 == 0 else WHITE_HEX)
    c.border = border_thin()
    c.alignment = left()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: Scored Portfolio
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Scored Portfolio")
ws2.sheet_view.showGridLines = False

display_cols = ['duration','credit_amount','age','purpose_label','employment_label',
                'checking_label','pd_score','risk_grade','lgd','ead','expected_loss','el_rate']
col_headers  = ['Duration (mo)','Credit Amount (DM)','Age','Purpose','Employment',
                'Checking Account','PD Score','Risk Grade','LGD','EAD (DM)',
                'Expected Loss (DM)','EL Rate']

col_widths = [13,18,8,15,14,16,10,10,8,14,18,10]
for i, w in enumerate(col_widths):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# Title
ws2.merge_cells(f'A1:{get_column_letter(len(display_cols))}1')
ws2['A1'] = 'SCORED LOAN PORTFOLIO — PD / LGD / EAD / Expected Loss'
ws2['A1'].font = font(bold=True, color=WHITE_HEX, size=12)
ws2['A1'].fill = fill(NAVY_HEX)
ws2['A1'].alignment = center()

# Headers
grade_colors_map = {
    'AAA': '1A7A4A','AA': '27AE60','A': '82E0AA',
    'BBB': GOLD_HEX,'BB': 'F39C12','B': 'E67E22','CCC': RED_HEX
}
for j, h in enumerate(col_headers):
    c = ws2.cell(row=2, column=j+1)
    c.value = h; c.font = font(bold=True, color=WHITE_HEX, size=9)
    c.fill = fill(BLUE_HEX); c.alignment = center(); c.border = border_thin()

# Data rows (first 200 for readability)
port_display = portfolio[display_cols].head(200)
for i, (_, row) in enumerate(port_display.iterrows()):
    r = i + 3
    bg = MGREY_HEX if i % 2 == 0 else WHITE_HEX
    for j, col in enumerate(display_cols):
        c = ws2.cell(row=r, column=j+1)
        val = row[col]
        if col == 'pd_score':
            c.value = round(float(val), 3)
            c.number_format = '0.0%'
        elif col in ['lgd','el_rate']:
            c.value = round(float(val), 3)
            c.number_format = '0.0%'
        elif col in ['credit_amount','ead','expected_loss']:
            c.value = int(val)
            c.number_format = '#,##0'
        else:
            c.value = val
        c.font = font(size=9)
        if col == 'risk_grade':
            grade_color = grade_colors_map.get(str(val), WHITE_HEX)
            c.fill = fill(grade_color)
            c.font = font(bold=True, color=WHITE_HEX, size=9)
        else:
            c.fill = fill(bg)
        c.border = border_thin()
        c.alignment = center()

ws2.freeze_panes = 'A3'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: Stress Test Results
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Stress Test")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 28
for col in ['B','C','D','E','F','G']:
    ws3.column_dimensions[col].width = 16

ws3.merge_cells('A1:G1')
ws3['A1'] = 'STRESS TEST & SCENARIO ANALYSIS — Basel II Framework'
ws3['A1'].font = font(bold=True, color=WHITE_HEX, size=12)
ws3['A1'].fill = fill(NAVY_HEX)
ws3['A1'].alignment = center()

stress_headers = ['Scenario','Avg PD','Avg EL Rate','Total EL (DM)','EL % Portfolio',
                  'vs Base (DM)','EL Increase %']
for j, h in enumerate(stress_headers):
    c = ws3.cell(row=2, column=j+1)
    c.value = h; c.font = font(bold=True, color=WHITE_HEX, size=9)
    c.fill = fill(BLUE_HEX); c.alignment = center(); c.border = border_thin()

stress_row_colors = [MGREY_HEX, 'FFE8E8', 'FFF5E0', 'F5E8FF', 'FFF0E0']
for i, (_, row) in enumerate(stress.iterrows()):
    r = i + 3
    bg = stress_row_colors[i]
    vals = [
        row['Scenario'], f"{row['Avg PD']:.1%}", f"{row['Avg EL Rate']:.1%}",
        int(row['Total EL (DM)']), f"{row['EL % Portfolio']:.1%}",
        int(row['vs Base (DM)']), f"{row['EL Increase %']:.0%}"
    ]
    for j, val in enumerate(vals):
        c = ws3.cell(row=r, column=j+1)
        c.value = val
        c.font = font(bold=(j==0), size=9)
        c.fill = fill(bg); c.border = border_thin()
        c.alignment = left() if j == 0 else center()
        if j in [3,5]:
            c.number_format = '#,##0'

# Grade migration
ws3.cell(row=10, column=1).value = 'GRADE MIGRATION MATRIX'
ws3.cell(row=10, column=1).font = font(bold=True, color=WHITE_HEX, size=11)
ws3.cell(row=10, column=1).fill = fill(NAVY_HEX)
ws3.merge_cells('A10:G10')
ws3['A10'].alignment = left()

grade_order = ['AAA','AA','A','BBB','BB','B','CCC']
scenarios_list = list(migration.columns)
for j, s in enumerate(scenarios_list):
    c = ws3.cell(row=11, column=j+2)
    c.value = s; c.font = font(bold=True, color=WHITE_HEX, size=8)
    c.fill = fill(BLUE_HEX); c.alignment = center(); c.border = border_thin()

grade_fill_colors = {
    'AAA':'1A7A4A','AA':'27AE60','A':'82E0AA',
    'BBB':GOLD_HEX,'BB':'F39C12','B':'E67E22','CCC':RED_HEX
}
for i, grade in enumerate(grade_order):
    r = 12 + i
    c = ws3.cell(row=r, column=1)
    c.value = grade
    c.font = font(bold=True, color=WHITE_HEX, size=9)
    c.fill = fill(grade_fill_colors[grade])
    c.border = border_thin(); c.alignment = center()
    for j, s in enumerate(scenarios_list):
        val = int(migration.loc[grade, s]) if grade in migration.index else 0
        cell = ws3.cell(row=r, column=j+2)
        cell.value = val
        cell.font = font(size=9)
        cell.fill = fill(MGREY_HEX if i % 2 == 0 else WHITE_HEX)
        cell.border = border_thin(); cell.alignment = center()
        cell.number_format = '#,##0'

# Capital adequacy
ws3.cell(row=21, column=1).value = 'CAPITAL ADEQUACY'
ws3.cell(row=21, column=1).font = font(bold=True, color=WHITE_HEX, size=11)
ws3.cell(row=21, column=1).fill = fill(NAVY_HEX)
ws3.merge_cells('A21:G21')
ws3['A21'].alignment = left()

total_exposure = 3271258
tier1 = total_exposure * 0.08
cap_headers = ['Scenario','Total EL (DM)','Tier 1 Capital (DM)','Buffer (DM)','Status']
for j, h in enumerate(cap_headers):
    c = ws3.cell(row=22, column=j+1)
    c.value = h; c.font = font(bold=True, color=WHITE_HEX, size=9)
    c.fill = fill(BLUE_HEX); c.alignment = center(); c.border = border_thin()

for i, (_, row) in enumerate(stress.iterrows()):
    r = 23 + i
    el = row['Total EL (DM)']
    buffer = tier1 - el
    status = '✅ Adequate' if buffer > 0 else '❌ Insufficient'
    bg = 'E8F8EE' if buffer > 0 else 'FFE8E8'
    for j, val in enumerate([row['Scenario'], int(el), int(tier1), int(buffer), status]):
        c = ws3.cell(row=r, column=j+1)
        c.value = val
        c.font = font(size=9, color=GREEN_HEX if status=='✅ Adequate' else RED_HEX if j==4 else '000000')
        c.fill = fill(bg); c.border = border_thin()
        c.alignment = left() if j in [0,4] else center()
        if j in [1,2,3]:
            c.number_format = '#,##0'

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: Model Performance
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Model Performance")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 35
ws4.column_dimensions['B'].width = 20
ws4.column_dimensions['C'].width = 35

ws4.merge_cells('A1:C1')
ws4['A1'] = 'PD MODEL PERFORMANCE — Logistic Regression with WoE Features'
ws4['A1'].font = font(bold=True, color=WHITE_HEX, size=12)
ws4['A1'].fill = fill(NAVY_HEX)
ws4['A1'].alignment = center()

metrics = [
    ('DISCRIMINATORY POWER', '', ''),
    ('AUC-ROC', '0.833', 'Excellent (benchmark: >0.70)'),
    ('Gini Coefficient', '0.666', 'Strong (benchmark: >0.40)'),
    ('5-Fold CV AUC', '0.806 ± 0.019', 'Stable — no overfitting'),
    ('', '', ''),
    ('CALIBRATION', '', ''),
    ('Brier Score', '0.169', 'Well calibrated (random = 0.250)'),
    ('', '', ''),
    ('CLASSIFICATION (threshold = 0.5)', '', ''),
    ('Overall Accuracy', '78%', ''),
    ('No Default Precision', '92%', 'Low false positives'),
    ('Default Recall', '85%', 'Captures 85% of actual defaults'),
    ('', '', ''),
    ('KEY PREDICTORS (Information Value)', '', ''),
    ('Checking Account Status', '0.666', 'Very Strong'),
    ('Credit History', '0.293', 'Medium'),
    ('Savings Account', '0.196', 'Medium'),
    ('Loan Purpose', '0.169', 'Medium'),
    ('Property Ownership', '0.113', 'Medium'),
    ('Employment Duration', '0.086', 'Weak'),
    ('Job Type', '0.009', 'Useless — excluded'),
    ('Telephone', '0.006', 'Useless — excluded'),
]

for i, (label, value, note) in enumerate(metrics):
    r = i + 2
    ws4.row_dimensions[r].height = 18
    is_header = (value == '' and label != '')
    for j, val in enumerate([label, value, note]):
        c = ws4.cell(row=r, column=j+1)
        c.value = val
        if is_header:
            c.font = font(bold=True, color=WHITE_HEX, size=9)
            c.fill = fill(BLUE_HEX)
        else:
            c.font = font(size=9, bold=(j==0))
            c.fill = fill(MGREY_HEX if i % 2 == 0 else WHITE_HEX)
        c.border = border_thin()
        c.alignment = left()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: Charts
# ══════════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Charts")
ws5.sheet_view.showGridLines = False
ws5.merge_cells('A1:P1')
ws5['A1'] = 'VISUALISATIONS — EDA · PD Model · Stress Test'
ws5['A1'].font = font(bold=True, color=WHITE_HEX, size=12)
ws5['A1'].fill = fill(NAVY_HEX)
ws5['A1'].alignment = center()

fig_files = ['fig1_eda.png', 'fig2_pd_model.png', 'fig3_stress_test.png']
fig_titles = ['Figure 1 — Exploratory Data Analysis',
              'Figure 2 — PD Model Performance & Risk Grades',
              'Figure 3 — Stress Test & Scenario Analysis']
fig_rows   = [3, 30, 57]

for fig_file, fig_title, start_row in zip(fig_files, fig_titles, fig_rows):
    path = BASE + fig_file
    if os.path.exists(path):
        ws5.cell(row=start_row-1, column=1).value = fig_title
        ws5.cell(row=start_row-1, column=1).font = font(bold=True, color=NAVY_HEX, size=11)
        img = XLImage(path)
        img.width  = 900
        img.height = 430
        ws5.add_image(img, f'A{start_row}')
        print(f"  Embedded: {fig_file}")
    else:
        print(f"  ⚠️ Not found: {fig_file}")

# ── Save ───────────────────────────────────────────────────────────────────────
output_path = BASE + 'Credit_Risk_Stress_Testing_Model.xlsx'
wb.save(output_path)
print(f"\n✅ Excel report saved: {output_path}")
print("   Sheets: Executive Summary · Scored Portfolio · Stress Test · Model Performance · Charts")
print("\nPhase 4 complete!")